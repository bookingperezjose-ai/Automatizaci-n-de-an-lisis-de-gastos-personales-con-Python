"""
Genera expenses_report.xlsx con:
- Hoja Dashboard: KPIs destacados + 4 gráficos
- Hoja Top10: top 10 transacciones + gráfico de barras
- Hoja Tabla_Dinamica: gasto por categoría x mes (con fórmulas)
- Hoja Outliers: transacciones anómalas detectadas
- Hoja Datos: transacciones limpias completas
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, PieChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Estilos ---
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FONT_KPI_LABEL = Font(name="Arial", size=10, bold=True, color="595959")
FONT_KPI_VALUE = Font(name="Arial", size=16, bold=True, color="1F4E78")
FONT_NORMAL = Font(name="Arial", size=10)
FILL_HEADER = PatternFill("solid", start_color="1F4E78")
FILL_TITLE = PatternFill("solid", start_color="1F4E78")
FILL_KPI = PatternFill("solid", start_color="EAF1F8")
FILL_OUTLIER = PatternFill("solid", start_color="FCE4E4")
THIN_BORDER = Border(*[Side(style="thin", color="B7C6D9")] * 4)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

CATEGORIAS = ["Alimentación", "Transporte", "Entretenimiento", "Servicios", "Otros"]


def estilizar_header(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def autofit(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def main():
    gastos = pd.read_csv("expenses_clean.csv", encoding="utf-8-sig", parse_dates=["Fecha"])
    mensual = pd.read_csv("monthly_summary.csv", encoding="utf-8-sig")
    outliers = pd.read_csv("outliers_detected.csv", encoding="utf-8-sig", parse_dates=["Fecha"])
    top10 = pd.read_csv("top10_expenses.csv", encoding="utf-8-sig", parse_dates=["Fecha"])
    kpis = pd.read_csv("kpis_summary.csv", index_col=0)

    meses = sorted(gastos["Mes"].unique())

    wb = Workbook()

    # ============ HOJA 1: DASHBOARD ============
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    ws["A1"] = "DASHBOARD — ANÁLISIS DE GASTOS PERSONALES (6 MESES)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_TITLE
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # --- KPIs destacados (fila 3-4) ---
    kpi_labels = [
        ("Gasto Total (6m)", "Gasto_Total_6meses", "R$"),
        ("Promedio Mensual", "Gasto_Promedio_Mensual", "R$"),
        ("Gasto Máximo", "Gasto_Máximo_Transacción", "R$"),
        ("Gasto Mínimo", "Gasto_Mínimo_Transacción", "R$"),
        ("Outliers Detectados", "N_Outliers_Detectados", ""),
        ("Proyección Próx. Mes", "Proyección_Próximo_Mes", "R$"),
    ]
    col = 1
    for label, key, prefix in kpi_labels:
        ws.cell(row=3, column=col, value=label).font = FONT_KPI_LABEL
        ws.cell(row=3, column=col).fill = FILL_KPI
        ws.cell(row=3, column=col).alignment = ALIGN_CENTER
        ws.cell(row=3, column=col).border = THIN_BORDER
        val = kpis.loc[key, "Valor"]
        val_str = f"{prefix} {val:,.2f}" if prefix else f"{int(val)}"
        ws.cell(row=4, column=col, value=val_str).font = FONT_KPI_VALUE
        ws.cell(row=4, column=col).fill = FILL_KPI
        ws.cell(row=4, column=col).alignment = ALIGN_CENTER
        ws.cell(row=4, column=col).border = THIN_BORDER
        col += 1
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 24

    # --- Tabla base para gráfico de evolución mensual (con fórmulas) ---
    r0 = 7
    ws.cell(row=r0, column=1, value="Mes").font = FONT_HEADER
    ws.cell(row=r0, column=2, value="Gasto Total").font = FONT_HEADER
    estilizar_header(ws, r0, 2)
    for i, mes in enumerate(meses):
        ws.cell(row=r0 + 1 + i, column=1, value=mes)
        total = gastos[gastos["Mes"] == mes]["Monto"].sum()
        ws.cell(row=r0 + 1 + i, column=2, value=round(total, 2)).number_format = '"R$" #,##0.00'
    r_evol_end = r0 + len(meses)

    # --- Tabla base para pie chart (categoría) ---
    r1 = r_evol_end + 3
    ws.cell(row=r1, column=1, value="Categoría").font = FONT_HEADER
    ws.cell(row=r1, column=2, value="Total Gastado").font = FONT_HEADER
    estilizar_header(ws, r1, 2)
    for i, cat in enumerate(CATEGORIAS):
        ws.cell(row=r1 + 1 + i, column=1, value=cat)
        total_cat = gastos[gastos["Categoría"] == cat]["Monto"].sum()
        ws.cell(row=r1 + 1 + i, column=2, value=round(total_cat, 2)).number_format = '"R$" #,##0.00'
    r_cat_end = r1 + len(CATEGORIAS)

    # --- Tabla variación mes a mes ---
    r2 = r_cat_end + 3
    ws.cell(row=r2, column=1, value="Mes").font = FONT_HEADER
    ws.cell(row=r2, column=2, value="Variación %").font = FONT_HEADER
    estilizar_header(ws, r2, 2)
    for i, mes in enumerate(meses):
        ws.cell(row=r2 + 1 + i, column=1, value=mes)
        fila_excel_actual = r0 + 1 + i
        if i == 0:
            ws.cell(row=r2 + 1 + i, column=2, value=0).number_format = '0.0%'
        else:
            fila_anterior = r0 + i
            formula = f'=IF(B{fila_anterior}=0,"",(B{fila_excel_actual}-B{fila_anterior})/B{fila_anterior})'
            ws.cell(row=r2 + 1 + i, column=2, value=formula).number_format = '0.0%'
    r_var_end = r2 + len(meses)

    autofit(ws, {"A": 22, "B": 18, "C": 18, "D": 18, "E": 20, "F": 20, "G": 18})

    # --- Gráfico 1: Línea evolución mensual ---
    chart1 = LineChart()
    chart1.title = "Evolución de Gasto Mensual"
    chart1.style = 12
    chart1.y_axis.title = "R$"
    chart1.x_axis.title = "Mes"
    data = Reference(ws, min_col=2, min_row=r0, max_row=r_evol_end)
    cats = Reference(ws, min_col=1, min_row=r0 + 1, max_row=r_evol_end)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.width, chart1.height = 14, 8
    ws.add_chart(chart1, "D7")

    # --- Gráfico 2: Pie chart categorías ---
    chart2 = PieChart()
    chart2.title = "Distribución de Gasto por Categoría"
    data = Reference(ws, min_col=2, min_row=r1, max_row=r_cat_end)
    cats = Reference(ws, min_col=1, min_row=r1 + 1, max_row=r_cat_end)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.width, chart2.height = 14, 8
    ws.add_chart(chart2, "D24")

    # --- Gráfico 3: Barras variación mes a mes ---
    chart3 = BarChart()
    chart3.title = "Variación de Gasto Mes a Mes (%)"
    chart3.style = 10
    chart3.y_axis.title = "% Variación"
    data = Reference(ws, min_col=2, min_row=r2, max_row=r_var_end)
    cats = Reference(ws, min_col=1, min_row=r2 + 1, max_row=r_var_end)
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.width, chart3.height = 14, 8
    ws.add_chart(chart3, "D41")

    # ============ HOJA 2: TOP 10 GASTOS ============
    ws2 = wb.create_sheet("Top10_Gastos")
    ws2.sheet_view.showGridLines = False
    headers = ["Fecha", "Descripción", "Categoría", "Monto"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    estilizar_header(ws2, 1, len(headers))
    for i, row in top10.iterrows():
        ws2.cell(row=i + 2, column=1, value=row["Fecha"].strftime("%Y-%m-%d"))
        ws2.cell(row=i + 2, column=2, value=row["Descripción"])
        ws2.cell(row=i + 2, column=3, value=row["Categoría"])
        ws2.cell(row=i + 2, column=4, value=round(row["Monto"], 2)).number_format = '"R$" #,##0.00'
    autofit(ws2, {"A": 14, "B": 32, "C": 18, "D": 14})

    chart4 = BarChart()
    chart4.title = "Top 10 Transacciones Más Altas"
    chart4.style = 11
    chart4.y_axis.title = "R$"
    data = Reference(ws2, min_col=4, min_row=1, max_row=11)
    cats = Reference(ws2, min_col=2, min_row=2, max_row=11)
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.width, chart4.height = 16, 9
    ws2.add_chart(chart4, "F1")

    # ============ HOJA 3: TABLA DINÁMICA (Categoría x Mes) ============
    ws3 = wb.create_sheet("Tabla_Dinamica")
    ws3.sheet_view.showGridLines = False
    ws3.cell(row=1, column=1, value="Categoría")
    for c, mes in enumerate(meses, 2):
        ws3.cell(row=1, column=c, value=mes)
    ws3.cell(row=1, column=len(meses) + 2, value="Total")
    estilizar_header(ws3, 1, len(meses) + 2)

    pivot = gastos.pivot_table(index="Categoría", columns="Mes", values="Monto", aggfunc="sum", fill_value=0)
    pivot = pivot.reindex(CATEGORIAS).fillna(0)

    for r, cat in enumerate(CATEGORIAS, 2):
        ws3.cell(row=r, column=1, value=cat)
        for c, mes in enumerate(meses, 2):
            val = round(pivot.loc[cat, mes], 2) if mes in pivot.columns else 0
            ws3.cell(row=r, column=c, value=val).number_format = '"R$" #,##0.00'
        col_letra_ini = get_column_letter(2)
        col_letra_fin = get_column_letter(len(meses) + 1)
        ws3.cell(row=r, column=len(meses) + 2,
                 value=f"=SUM({col_letra_ini}{r}:{col_letra_fin}{r})").number_format = '"R$" #,##0.00'

    r_total = len(CATEGORIAS) + 2
    ws3.cell(row=r_total, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, len(meses) + 3):
        col_letra = get_column_letter(c)
        ws3.cell(row=r_total, column=c,
                 value=f"=SUM({col_letra}2:{col_letra}{r_total-1})").number_format = '"R$" #,##0.00'
        ws3.cell(row=r_total, column=c).font = Font(bold=True)
    autofit(ws3, {get_column_letter(i): 16 for i in range(1, len(meses) + 3)})

    # ============ HOJA 4: OUTLIERS DETECTADOS ============
    ws4 = wb.create_sheet("Outliers_Detectados")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:E1")
    ws4["A1"] = f"Transacciones anómalas (Monto > Q3 + 1.5×IQR = R$ {kpis.loc['Umbral_Outlier (Q3+1.5*IQR)','Valor']:,.2f})"
    ws4["A1"].font = Font(name="Arial", size=11, bold=True, italic=True, color="C00000")
    headers = ["Fecha", "Descripción", "Categoría", "Monto", "Umbral Aplicado"]
    for c, h in enumerate(headers, 1):
        ws4.cell(row=3, column=c, value=h)
    estilizar_header(ws4, 3, len(headers))
    for i, row in outliers.iterrows():
        r = i + 4
        ws4.cell(row=r, column=1, value=row["Fecha"].strftime("%Y-%m-%d"))
        ws4.cell(row=r, column=2, value=row["Descripción"])
        ws4.cell(row=r, column=3, value=row["Categoría"])
        ws4.cell(row=r, column=4, value=round(row["Monto"], 2)).number_format = '"R$" #,##0.00'
        ws4.cell(row=r, column=5, value=round(row["Umbral_aplicado"], 2)).number_format = '"R$" #,##0.00'
        for c in range(1, 6):
            ws4.cell(row=r, column=c).fill = FILL_OUTLIER
    autofit(ws4, {"A": 14, "B": 34, "C": 18, "D": 14, "E": 16})

    # ============ HOJA 5: DATOS COMPLETOS ============
    ws5 = wb.create_sheet("Datos_Completos")
    ws5.sheet_view.showGridLines = False
    cols = ["Fecha", "Descripción", "Categoría", "Monto", "Tipo", "Saldo anterior", "Saldo posterior"]
    for c, h in enumerate(cols, 1):
        ws5.cell(row=1, column=c, value=h)
    estilizar_header(ws5, 1, len(cols))
    for i, row in gastos.iterrows():
        r = i + 2
        ws5.cell(row=r, column=1, value=row["Fecha"].strftime("%Y-%m-%d"))
        ws5.cell(row=r, column=2, value=row["Descripción"])
        ws5.cell(row=r, column=3, value=row["Categoría"])
        ws5.cell(row=r, column=4, value=round(row["Monto"], 2)).number_format = '"R$" #,##0.00'
        ws5.cell(row=r, column=5, value=row["Tipo"])
        ws5.cell(row=r, column=6, value=round(row["Saldo anterior"], 2)).number_format = '"R$" #,##0.00'
        ws5.cell(row=r, column=7, value=round(row["Saldo posterior"], 2)).number_format = '"R$" #,##0.00'
    autofit(ws5, {"A": 14, "B": 32, "C": 18, "D": 14, "E": 12, "F": 16, "G": 16})

    wb.save("expenses_report.xlsx")
    print("expenses_report.xlsx generado con 5 hojas y 4 gráficos.")


if __name__ == "__main__":
    main()
